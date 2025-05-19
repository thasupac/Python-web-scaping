# from bs4 import BeautifulSoup

# import requests

# heml_text = requests.get('https://www.timesjobs.com/candidate/job-search.html?searchType=personalizedSearch&from=submit&searchTextSrc=&searchTextText=&txtKeywords=python&txtLocation=')

# print(heml_text)

# soup = BeautifulSoup(heml_text, 'lxml')
# print(soup)
# job = soup.find('li', class_ = 'clearfix job-bx wht-shd-bx')
# # print(jobs)


from bs4 import BeautifulSoup
import requests
from openpyxl import load_workbook

url = 'https://www.timesjobs.com/candidate/job-search.html?searchType=personalizedSearch&from=submit&searchTextSrc=&searchTextText=&txtKeywords=python&txtLocation='
response = requests.get(url)
#print(response)
html = response.content

soup = BeautifulSoup(html, "html.parser")
jobs = soup.find_all('li', class_ = 'clearfix job-bx wht-shd-bx')

# jobs_list = {'company':[],'skills':[],'publish':[]}
jobs_list = []

for job in jobs:

    #companyname
    company_name_split = []
    company_name = job.find('h3', class_ = 'joblist-comp-name').text.replace(' ','')
    for name in company_name.split('\n'):
        stripped_name = name.strip()
        if stripped_name:
            company_name_split.append(stripped_name)

    if company_name_split:
        first_company_name = company_name_split[0]
    else:
        first_company_name = ""

    # jobs_list['company'].append(first_company_name)

    #skill
    skill_split = []
    skills = job.find('span', class_ = 'srp-skills').text.replace(' ','')
    for sum_skills in skills.split('\n'):
        skill_stripped = sum_skills.strip()
        if skill_stripped:
            skill_split.append(skill_stripped)
    
    if skill_split:
        skill_sum = skill_split[0]
    else:
        skill_sum = ""

    # jobs_list['skills'].append(skill_sum)

    #pubblish
    publish_split = []
    publish_date = job.find('span', class_ = 'sim-posted').span.text
    for publish_sum in publish_date.split('\n'):
        publish_stripped = publish_sum.strip()
        if publish_stripped:
            publish_split.append(publish_stripped)
    
    if publish_split:
        sum_publish = publish_split[0]
    else:
        sum_publish =""

    jobs_list.append([first_company_name,skill_sum,sum_publish])

   
row_product = 1
for comp in jobs_list:
    excel_file = load_workbook(filename='web_scraping.xlsx')
    sheet = excel_file.active


    sheet.cell(row=row_product, column=1).value = comp[0]
    sheet.cell(row=row_product, column=2).value = comp[1]
    sheet.cell(row=row_product, column=3).value = comp[2]

    row_product = row_product+1

    excel_file.save('web_scraping.xlsx')

    print(comp)





